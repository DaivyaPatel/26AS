import pandas as pd
import pdfplumber
import os
import re
from fuzzywuzzy import fuzz
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─────────────────────────────────────────
# 1. FORM 26AS PARSER
# ─────────────────────────────────────────

def parse_26as(filepath):
    ext = os.path.splitext(filepath)[1].lower()
    records = []

    if ext == ".pdf":
        with pdfplumber.open(filepath) as pdf:
            for page in pdf.pages:
                text = page.extract_text() or ""
                lines = text.split("\n")
                
                for line in lines:
                    line = line.strip()
                    if not line:
                        continue

                    # Targets: <Sr.No> <COMPANY NAME> <TAN> <AMT PAID> <TAX DEDUCTED> <TDS DEPOSITED>
                    match = re.search(r'^\d+\s+(.+?)\s+([A-Z]{4}[0-9]{5}[A-Z])\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})', line)
                    if match:
                        name = match.group(1).strip()
                        tan = match.group(2).strip()
                        tds_deposited = float(match.group(5).replace(",", ""))
                        
                        if tds_deposited > 0:
                            records.append({
                                "deductor_name": name,
                                "tan": tan,
                                "tds_amount": tds_deposited
                            })

    df_result = pd.DataFrame(records)
    if not df_result.empty:
        df_result = df_result.groupby(["deductor_name", "tan"]).agg({"tds_amount": "sum"}).reset_index()
    return df_result


# ─────────────────────────────────────────
# 2. CONTINUOUS LAYOUT TALLY PARSER
# ─────────────────────────────────────────

def parse_tally_ledger(filepath, ledger_type="direct"):
    ext = os.path.splitext(filepath)[1].lower()
    records = []

    if ext != ".pdf":
        return pd.DataFrame()

    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            for raw_line in text.split("\n"):
                line = raw_line.strip()
                if not line:
                    continue

                line_lower = line.lower()

                # SKIP opening balance, credits, and non-TDS lines
                if any(x in line_lower for x in [
                    "opening balance", "closing balance", "brought forward",
                    "carried over", "continued", "ledger account", "page ",
                    "particulars", "date particulars", "udyam reg no",
                    "a t m s & co", "suhas shinde",  # skip capital transfers
                    "tax a/c",  # skip tax payment credits
                ]):
                    continue

                # Only capture "To" lines (debit = TDS received)
                if not re.search(r'\bTo\b', line, re.IGNORECASE):
                    continue

                # Skip "By" lines — those are credits (payments OUT)
                if re.search(r'\bBy\b', line, re.IGNORECASE):
                    continue

                amount_match = re.search(r'([\d,]+\.\d{2})$', line)
                if not amount_match:
                    continue

                try:
                    amount = float(amount_match.group(1).replace(",", ""))
                except:
                    continue

                if amount <= 0:
                    continue

                # Strip voucher type and number
                party_part = re.sub(
                    r'(Journal|Receipt|Payment|Contra|Sales|Purchase)\s+\d+\s+[\d,]+\.\d{2}$',
                    '', line, flags=re.IGNORECASE
                )
                party_part = re.sub(r'^\d{1,2}-[A-Za-z]{3}-\d{2,4}\s+', '', party_part)
                party_part = re.sub(r'^To\s+', '', party_part, flags=re.IGNORECASE)
                party_name = party_part.strip()

                if len(party_name) < 3:
                    continue

                records.append({
                    "party_name": party_name,
                    "tds_amount": amount,
                    "ledger_type": ledger_type
                })

    df = pd.DataFrame(records)
    if not df.empty:
        df = df.groupby(["party_name", "ledger_type"], as_index=False)["tds_amount"].sum()
    return df

# ─────────────────────────────────────────
# 3. HIGH-SPEED CROSS-MATCHING ENGINE
# ─────────────────────────────────────────

def normalize_name(name):
    name = str(name).upper()

    remove_words = [
        "PRIVATE LIMITED",
        "PVT LTD",
        "LIMITED",
        "LTD",
        "LLP",
        "CO",
        "& CO",
        "SERVICES",
        "SERVICE",
        "SOLUTIONS",
        "SOLUTION",
        "TECHNOLOGIES",
        "TECHNOLOGY",
        "INDIA",
        "PVT",
        "PRIVATE"
    ]

    for word in remove_words:
        name = name.replace(word, " ")

    name = re.sub(r'[^A-Z0-9 ]', ' ', name)
    name = re.sub(r'\s+', ' ', name)

    return name.strip()


def match_entries(df_26as, df_direct, df_indirect):
    df_tally = pd.concat([df_direct, df_indirect], ignore_index=True) if (not df_direct.empty or not df_indirect.empty) else pd.DataFrame()
    results = []
    used_tally_idx = set()

    if df_26as.empty:
        return results

    tally_pool = []
    if not df_tally.empty:
        for idx, row in df_tally.iterrows():
            tally_pool.append({
                "idx": idx,
                "party_name": row["party_name"],
                "norm_name": normalize_name(row["party_name"]),
                "tds_amount": row["tds_amount"],
                "ledger_type": row["ledger_type"]
            })

    for _, row_26as in df_26as.iterrows():
        n_26as = row_26as["deductor_name"]
        t_26as = row_26as["tan"]
        a_26as = row_26as["tds_amount"]
        norm_26as = normalize_name(n_26as)

        best_match = None
        best_score = 0
        best_idx = None

        # Exact Match Pass
        for item in tally_pool:
            if item["idx"] in used_tally_idx:
                continue
            if norm_26as == item["norm_name"]:
                best_match = item
                best_score = 100
                best_idx = item["idx"]
                break

        # Fuzzy Match Pass
        if best_match is None and tally_pool:
            for item in tally_pool:
                if item["idx"] in used_tally_idx:
                    continue

                score = fuzz.token_set_ratio(norm_26as, item["norm_name"])
                if abs(a_26as - item["tds_amount"]) < 5.0:
                    score += 25

                if score > best_score:
                    best_score = score
                    best_match = item
                    best_idx = item["idx"]

        if best_match is not None and best_score >= 60:
            used_tally_idx.add(best_idx)
            a_tally = best_match["tds_amount"]
            diff = round(a_26as - a_tally, 2)
            
            status = "MATCHED" if abs(diff) <= 10.0 else "PARTIAL"
            reason = "Reconciliation Match Verified" if status == "MATCHED" else f"Amount Mismatch: ₹{abs(diff):,.2f}"

            results.append({
                "status": status, "reason": reason, "deductor_26as": n_26as, "tan_26as": t_26as,
                "amount_26as": a_26as, "party_tally": best_match["party_name"], "amount_tally": a_tally,
                "ledger_type": best_match["ledger_type"], "amount_diff": diff, "name_score": min(best_score, 100)
            })
        else:
            results.append({
                "status": "UNMATCHED", "reason": "No match found in Tally books",
                "deductor_26as": n_26as, "tan_26as": t_26as, "amount_26as": a_26as,
                "party_tally": "", "amount_tally": 0, "ledger_type": "", "amount_diff": a_26as, "name_score": best_score
            })

    if not df_tally.empty:
        for item in tally_pool:
            if item["idx"] not in used_tally_idx:
                results.append({
                    "status": "TALLY_ONLY", "reason": "Present in Tally books but missing from Form 26AS",
                    "deductor_26as": "", "tan_26as": "", "amount_26as": 0,
                    "party_tally": item["party_name"], "amount_tally": item["tds_amount"],
                    "ledger_type": item["ledger_type"], "amount_diff": -item["tds_amount"], "name_score": 0
                })

    return results


def generate_report(results, output_folder):
    wb = Workbook()
    ws = wb.active
    ws.title = "Full Reconciliation"
    
    headers = ["Status", "Reason", "Deductor Name (26AS)", "TAN (26AS)", "Amount in 26AS (₹)", "Party Name (Tally)", "Ledger Type", "Amount in Tally (₹)", "Difference (₹)", "Match Score %"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1F4E79")
    white_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    for col_num in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = white_font
        cell.border = thin_border

    status_colors = {"MATCHED": "C6EFCE", "PARTIAL": "FFEB9C", "UNMATCHED": "FFC7CE", "TALLY_ONLY": "BDD7EE"}

    for r in results:
        ws.append([r["status"], r["reason"], r["deductor_26as"], r["tan_26as"], r["amount_26as"], r["party_tally"], r["ledger_type"], r["amount_tally"], r["amount_diff"], r["name_score"]])
        current_fill = PatternFill("solid", fgColor=status_colors.get(r["status"], "FFFFFF"))
        for col_num in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.fill = current_fill
            cell.border = thin_border

    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    output_path = os.path.join(output_folder, "reconciliation_report.xlsx")
    wb.save(output_path)
    return output_path


def reconcile_files(path_26as, path_direct, path_indirect, output_folder):

    # Parse files
    df_26as = parse_26as(path_26as)
    df_direct = parse_tally_ledger(path_direct, "direct")
    df_indirect = parse_tally_ledger(path_indirect, "indirect")
    print("\n===== DIRECT LEDGER SAMPLE =====")
    print(df_direct.head(20))

    print("\n===== INDIRECT LEDGER SAMPLE =====")
    print(df_indirect.head(20))

    print("\n===== TOTALS =====")
    print("DT TOTAL =", df_direct["tds_amount"].sum())
    print("IDT TOTAL =", df_indirect["tds_amount"].sum())

    # Validation
    if df_26as.empty and df_direct.empty and df_indirect.empty:
        raise ValueError(
            "Critical Layout Error: Unable to extract data fields from source files."
        )

    # Debug Totals
    print("\n========== EXTRACTION SUMMARY ==========")

    total_26as = (
        df_26as["tds_amount"].sum()
        if not df_26as.empty
        else 0
    )

    total_direct = (
        df_direct["tds_amount"].sum()
        if not df_direct.empty
        else 0
    )

    total_indirect = (
        df_indirect["tds_amount"].sum()
        if not df_indirect.empty
        else 0
    )

    total_tally = total_direct + total_indirect

    print(f"26AS TOTAL   = ₹{total_26as:,.2f}")
    print(f"DT TOTAL     = ₹{total_direct:,.2f}")
    print(f"IDT TOTAL    = ₹{total_indirect:,.2f}")
    print(f"TALLY TOTAL  = ₹{total_tally:,.2f}")
    print("========================================\n")

    # Reconciliation
    results = match_entries(
        df_26as,
        df_direct,
        df_indirect
    )

    # Generate Excel Report
    output_path = generate_report(
        results,
        output_folder
    )

    # Summary Statistics
    summary = {
        "matched": sum(
            1 for r in results
            if r["status"] == "MATCHED"
        ),

        "partial": sum(
            1 for r in results
            if r["status"] == "PARTIAL"
        ),

        "unmatched": sum(
            1 for r in results
            if r["status"] == "UNMATCHED"
        ),

        "tally_only": sum(
            1 for r in results
            if r["status"] == "TALLY_ONLY"
        ),

        "total_26as": round(
            total_26as,
            2
        ),

        "total_tally": round(
            total_tally,
            2
        ),

        "net_difference": round(
            total_26as - total_tally,
            2
        )
    }

    return output_path, summary